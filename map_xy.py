import os
import numpy as np
import motor_controller
import hp_line_scan
from adlink_card import EncoderException
from datetime import datetime
import openpyxl
import openpyxl.utils
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment, Border, Side
import colorcet  # for conditional formatting: colour scales

# Metadata for the file
magnet = 'PITZ Compensation Solenoid'
current = 30
comment = 'XZ scan at Y centre -10mm'
# Define the scan axes and range
# Axis 1 is the primary axis, this will use line_scan with on-the-fly scanning if possible
axis1, ax1_start, ax1_stop, ax1_step = 'z', -60, 90, 0.5
axis2, ax2_start, ax2_stop, ax2_step = 'x', -0.5, 10, 0.5
axis3, ax3_pos = 'y', 2.579 + 10
assert len(set((axis1, axis2, axis3))) == 3  # check we haven't tried to use the same axis twice

# What type of file to save?
save_csv = True
save_xlsx = True
# Where to save the file(s)?
path = r'\\fed.cclrc.ac.uk\Org\NLab\ASTeC\Apsv4\Astec\IDs and Magnets\Data\PITZ solenoids\Compensation solenoid 15050'
basename = '14 xz scan (y centre +10mm)'

# Set up the scan
mc = motor_controller.MotorController()
ax2_values = np.linspace(ax2_start, ax2_stop, int((ax2_stop - ax2_start) / ax2_step + 1))
line_scan = hp_line_scan.LineScan(axis1, ax1_start, ax1_stop, ax1_step, mc=mc)
hp = line_scan.hp
field_units = 'mT'
hp.setUnits(field_units)

csv_filename = os.path.join(path, basename + '.csv')
xlsx_filename = os.path.join(path, basename + '.xlsx')

# Produce a header for the file(s)
header = (('Date/time', datetime.now().strftime('%d/%m/%y %H:%M:%S')),
          ('Magnet under test', magnet),
          ('Magnet current [A]', current),
          ('Probe', f'{hp.probe.manufacturer_name} {hp.probe.model_name} S/N {hp.probe.serial_number}'),
          ('Averages', hp.getAverages()),
          ('Probe range', hp.getRange()[1]),
          ('Comment', comment),
          (f'{axis3} position [mm]', ax3_pos),
          )

if save_xlsx:
    # Create an Excel file and start writing data to it
    # Warning: this will overwrite an existing file without asking
    workbook = openpyxl.Workbook()
    # Create the 'Info' tab containing the metadata
    info_sheet = workbook['Sheet']  # automatically-named first worksheet
    info_sheet.title = "Info"
    info_sheet.column_dimensions['A'].width = 20.0  # header attribute names
    info_sheet.column_dimensions['B'].width = 50.0  # header attribute values
    for row, (name, value) in enumerate(header):
        info_sheet.cell(row + 1, 1, value=name).font = Font(bold=True)
        info_sheet.cell(row + 1, 2, value=value)
    # Create a sheet for each field direction. Each sheet will have a grid of XY points.
    field_dirs = ('Bx', 'By', 'Bz')
    field_sheets = [workbook.create_sheet(name) for name in field_dirs]
    array_range = 'C3:{}{}'.format(openpyxl.utils.get_column_letter(len(line_scan.pos_values) + 3), len(ax2_values) + 3)
    # Generate some nice conditional formatting using Peter Kovesi's colour maps
    # See https://peterkovesi.com/projects/colourmaps/ for more details
    scale = colorcet.blues  # light-blue colour scale - black text should be visible for all colours
    rule = ColorScaleRule(start_color=scale[0][1:], start_type='min',
                          mid_color=scale[128][1:], mid_type='percentile', mid_value=50,
                          end_color=scale[-1][1:], end_type='max')
    thin = Side(border_style="thin", color="000000")
    for sheet, name in zip(field_sheets, field_dirs):
        # Insert axis titles and axes into the sheet
        cell = sheet.cell(1, 1, value=f'{name} [{field_units}]')
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells('A1:B2')

        cell = sheet.cell(1, 3, value=f"{axis1} [mm]")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.merge_cells(start_row=1, end_row=1, start_column=3, end_column=len(line_scan.pos_values)+2)
        for row in range(3, len(line_scan.pos_values) + 3):
            sheet.cell(1, row).border = Border(bottom=thin)  # need to do every cell when merged

        cell = sheet.cell(3, 1, value=f"{axis2} [mm]")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", text_rotation=90)
        cell.border = Border(right=thin)
        sheet.column_dimensions['A'].width = 4
        sheet.merge_cells(start_column=1, end_column=1, start_row=3, end_row=len(ax2_values) + 2)
        for row in range(3, len(ax2_values) + 3):
            sheet.cell(row, 1).border = Border(right=thin)  # need to do every cell when merged

        [sheet.cell(row + 3, 2, value=y) for row, y in enumerate(ax2_values)]
        [sheet.cell(2, col + 3, value=x) for col, x in enumerate(line_scan.pos_values)]
        sheet.conditional_formatting.add(array_range, rule)
        for row in sheet[array_range]:
            for cell in row:
                cell.number_format = '0.000'
                sheet.column_dimensions[cell.column].width = 7  # seems OK to fit in 3 sig figs

if save_csv:
    # Write header and columns to CSV file
    csv_file = open(csv_filename, 'a')
    csv_file.writelines('{},{}\n'.format(*l) for l in header)
    csv_file.write(f'{axis1} [mm],{axis2} [mm],Bx [{field_units}],By [{field_units}],Bz [{field_units}]')

print(f'{axis3} = {ax3_pos} mm')
mc.axis[axis3].move(ax3_pos, wait=True)
start = datetime.now()
eta = None

# Run the scan, invoking line_scan for each position along axis 2
for i, y in enumerate(ax2_values):
    if i > 0:
        progress = i / len(ax2_values)
        elapsed = datetime.now() - start
        eta = start + elapsed / progress
    print(f'{axis2} = {y} mm' + (eta.strftime(', ETA %H:%M') if eta else ''))
    mc.axis[axis2].move(y, wait=True)
    retries = 0
    ok = False
    while not ok:
        try:
            line_scan.run()
            ok = True
        except (hp_line_scan.MissedTriggerError, EncoderException) as e:  # sometimes we get a little hiccup
            mc.axis[axis1].stop()
            retries += 1
            print(e)
            if retries % 5 == 0 and input(f'Scan failed after {retries} retries. Try again? [Y/n]').upper() in ('', 'Y'):
                raise  # break out
    # line_scan.field_values = np.random.rand(len(line_scan.pos_values), 3) - 0.5

    # Record the data in the file(s)
    for column, (x, field) in enumerate(zip(line_scan.pos_values, line_scan.field_values)):
        if save_xlsx:
            for sheet, field_component in zip(field_sheets, field):
                sheet.cell(i + 3, column + 3, value=field_component)
        if save_csv:
            csv_file.write(','.join([f'{p:.3f}' for p in np.concatenate([[x, y], field])]) + '\n')

    # Save as we go along in case of unforeseen errors
    if save_csv:
        csv_file.flush()
    if save_xlsx:
        workbook.save(xlsx_filename)

if save_csv:
    csv_file.close()
