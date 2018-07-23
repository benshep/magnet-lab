import openpyxl.utils
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment, Border, Side
import colorcet  # for conditional formatting: colour scales

# Create an Excel file and start writing data to it
# Warning: this will overwrite an existing file without asking
workbook = openpyxl.Workbook()
# Create the 'Info' tab containing the metadata
info_sheet = workbook['Sheet']  # automatically-named first worksheet
info_sheet.title = "Info"
info_sheet.column_dimensions['A'].width = 20.0  # header attribute names
info_sheet.column_dimensions['B'].width = 50.0  # header attribute values
for row, (name, value) in enumerate(header):
    info_sheet.cell(row + 1, 1, value=name).font = Font(bold=True)
    info_sheet.cell(row + 1, 2, value=value)
# Create a sheet for each field direction. Each sheet will have a grid of XY points.
field_dirs = ('Bx', 'By', 'Bz')
field_sheets = [workbook.create_sheet(name) for name in field_dirs]
array_range = 'C3:{}{}'.format(openpyxl.utils.get_column_letter(len(line_scan.pos_values) + 3), len(ax2_values) + 3)
# Generate some nice conditional formatting using Peter Kovesi's colour maps
# See https://peterkovesi.com/projects/colourmaps/ for more details
scale = colorcet.blues  # light-blue colour scale - black text should be visible for all colours
rule = ColorScaleRule(start_color=scale[0][1:], start_type='min',
                      mid_color=scale[128][1:], mid_type='percentile', mid_value=50,
                      end_color=scale[-1][1:], end_type='max')
thin = Side(border_style="thin", color="000000")
for sheet, name in zip(field_sheets, field_dirs):
    # Insert axis titles and axes into the sheet
    cell = sheet.cell(1, 1, value=f'{name} [{field_units}]')
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells('A1:B2')

    cell = sheet.cell(1, 3, value=f"{axis1} [mm]")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=1, end_row=1, start_column=3, end_column=len(line_scan.pos_values)+2)
    for row in range(3, len(line_scan.pos_values) + 3):
        sheet.cell(1, row).border = Border(bottom=thin)  # need to do every cell when merged

    cell = sheet.cell(3, 1, value=f"{axis2} [mm]")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(vertical="center", text_rotation=90)
    cell.border = Border(right=thin)
    sheet.column_dimensions['A'].width = 4
    sheet.merge_cells(start_column=1, end_column=1, start_row=3, end_row=len(ax2_values) + 2)
    for row in range(3, len(ax2_values) + 3):
        sheet.cell(row, 1).border = Border(right=thin)  # need to do every cell when merged

    [sheet.cell(row + 3, 2, value=y) for row, y in enumerate(ax2_values)]
    [sheet.cell(2, col + 3, value=x) for col, x in enumerate(line_scan.pos_values)]
    sheet.conditional_formatting.add(array_range, rule)
    for row in sheet[array_range]:
        for cell in row:
            cell.number_format = '0.000'
            sheet.column_dimensions[cell.column].width = 7  # seems OK to fit in 3 sig figs

# for each field value
    for sheet, field_component in zip(field_sheets, field):
        sheet.cell(i + 3, column + 3, value=field_component)


workbook.save(xlsx_filename)
